"""Generate detailed compliance report in Excel format."""
from __future__ import annotations

import argparse
import re
import xml.etree.ElementTree as ET
from typing import List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from utils import count_l1_l2

CHECK_PATTERN = re.compile(r"^(?P<check>\S+)\s+\((?P<level>L[12])\)\s+(?P<desc>.+)$")


def _strip_text(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    value = value.strip()
    return value if value else None


def _find_text(element: ET.Element, tag: str) -> Optional[str]:
    child = element.find(tag)
    if child is not None and child.text is not None:
        return _strip_text(child.text)

    for child in element:
        if child.tag.endswith(tag):
            if child.text is not None:
                return _strip_text(child.text)

    return None


def parse_check_name(check_name: Optional[str]) -> Tuple[str, str, str]:
    if not check_name:
        return "", "", ""

    match = CHECK_PATTERN.match(check_name)
    if match:
        return (
            match.group("check"),
            match.group("level"),
            match.group("desc"),
        )

    return "", "", check_name


def extract_items(scan_path: str) -> Tuple[Optional[str], List[dict]]:
    tree = ET.parse(scan_path)
    root = tree.getroot()

    host_ip: Optional[str] = None
    for report_host in root.iter("ReportHost"):
        host_ip = report_host.attrib.get("name") or _find_text(report_host, "host-ip")
        if host_ip:
            break

    items: List[dict] = []

    for report_item in root.iter("ReportItem"):
        compliance_text = _find_text(report_item, "compliance")
        compliance_result = _find_text(report_item, "compliance-result")
        if compliance_text and compliance_text.lower() != "true" and not compliance_result:
            continue
        if compliance_text is None and compliance_result is None:
            continue

        check_name = _find_text(report_item, "compliance-check-name")
        check_number, level, description = parse_check_name(check_name)

        info_text = _find_text(report_item, "compliance-info") or ""
        solution_text = _find_text(report_item, "compliance-solution") or ""
        impact_text = _find_text(report_item, "compliance-impact") or ""

        if not impact_text and solution_text:
            split_token = "Impact:"
            if split_token in solution_text:
                solution_part, impact_part = solution_text.split(split_token, 1)
                solution_text = solution_part.strip()
                impact_text = impact_part.strip()

        items.append(
            {
                "check_number": check_number,
                "level": level,
                "description": description,
                "result": compliance_result or "Unknown",
                "info": info_text,
                "solution": solution_text,
                "impact": impact_text,
                "policy_value": _find_text(report_item, "compliance-policy-value") or "",
                "actual_value": _find_text(report_item, "compliance-actual-value") or "",
            }
        )

    return host_ip, items


def build_excel(items, output_path, host_ip=None):
    """Build Excel workbook with compliance details for single file."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Compliance Details"

    # Define styles
    header_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    header_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Column widths
    column_widths = [15, 15, 12, 25, 15, 20, 20, 20, 20, 20]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[chr(64 + i)].width = width

    # Headers: IP Address, Check Number, Level, Description, Result, Info, Solution, Impact, Policy Value, Configured Value
    headers = [
        "IP Address",
        "Check Number",
        "Level",
        "Description",
        "Result",
        "Info",
        "Solution",
        "Impact",
        "Policy Value",
        "Configured Value"
    ]

    # Add header row
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Add data rows
    for row_num, item in enumerate(items, 2):
        row_data = [
            host_ip or "",
            item.get("check_number", ""),
            item.get("level", ""),
            item.get("description", ""),
            item.get("result", ""),
            item.get("info", ""),
            item.get("solution", ""),
            item.get("impact", ""),
            item.get("policy_value", ""),
            item.get("actual_value", ""),
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Freeze the header row
    worksheet.freeze_panes = "A2"

    workbook.save(output_path)


def build_excel_combined(all_files_data, output_path):
    """Build combined Excel workbook with all files, separated by blank rows."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Combined Compliance"

    # Define styles
    header_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    file_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Column widths
    column_widths = [15, 15, 12, 25, 15, 20, 20, 20, 20, 20]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[chr(64 + i)].width = width

    # Headers
    headers = [
        "IP Address",
        "Check Number",
        "Level",
        "Description",
        "Result",
        "Info",
        "Solution",
        "Impact",
        "Policy Value",
        "Configured Value"
    ]

    current_row = 1

    # Process each file's data
    for file_name, host_ip, items in all_files_data:
        # Add file name header (green)
        file_header_cell = worksheet.cell(row=current_row, column=1)
        file_header_cell.value = f"📄 File: {file_name}"
        file_header_cell.font = Font(bold=True, size=12, color="FFFFFF")
        file_header_cell.fill = file_header_fill
        
        for col in range(1, 11):
            cell = worksheet.cell(row=current_row, column=col)
            if col == 1:
                cell.value = f"📄 File: {file_name}"
                cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = file_header_fill
            cell.border = border
        
        current_row += 1

        # Add column headers (blue)
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=current_row, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        current_row += 1

        # Add data rows
        for item in items:
            row_data = [
                host_ip or "",
                item.get("check_number", ""),
                item.get("level", ""),
                item.get("description", ""),
                item.get("result", ""),
                item.get("info", ""),
                item.get("solution", ""),
                item.get("impact", ""),
                item.get("policy_value", ""),
                item.get("actual_value", ""),
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=current_row, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            current_row += 1

        # Add blank row separator
        current_row += 1

    workbook.save(output_path)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Create a detailed compliance report Excel from Nessus XML.")
    parser.add_argument("scan_path", help="Path to Nessus .nessus/.xml file")
    parser.add_argument(
        "--output",
        default="compliance_detailed_report.xlsx",
        help="Output .xlsx path (default: compliance_detailed_report.xlsx)",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    host_ip, items = extract_items(args.scan_path)
    build_excel(items, args.output, host_ip=host_ip)
    print(f"Wrote {len(items)} detailed compliance records to {args.output}")


if __name__ == "__main__":
    main()
